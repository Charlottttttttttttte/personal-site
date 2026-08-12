#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
临池录库存同步脚本
将 /Users/charlotte/Desktop/临池录-导入模板.xlsx 转换为
personal-site/data/shufa-inventory.json，供网站展示。

用法：
  python3 scripts/sync_inventory.py            # 仅生成 JSON
  python3 scripts/sync_inventory.py --open      # 生成后打开浏览器预览
"""
import csv, json, os, re, sys
from datetime import date

# 兼容两种读取：优先 openpyxl（.xlsx），失败回退到已导出的 CSV
XLSX = "/Users/charlotte/Desktop/临池录-导入模板.xlsx"
JSON_OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "shufa-inventory.json")


def read_rows():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [r for r in ws.iter_rows(values_only=True)
                if any(v is not None and str(v).strip() != "" for v in r)]
    except Exception as e:
        print(f"[警告] openpyxl 读取失败({e})，尝试 CSV 回退")
        # 从 xlsx 里找同目录 CSV（若用户手动导出过）
        base = os.path.splitext(XLSX)[0]
        for cand in [base + ".csv", XLSX.replace(".xlsx", ".csv")]:
            if os.path.exists(cand):
                with open(cand, encoding="utf-8-sig") as f:
                    return list(csv.reader(f))
        raise


def col_index(headers, *names):
    for i, h in enumerate(headers):
        hs = str(h or "").strip()
        for n in names:
            if n == hs or hs.startswith(n) or n in hs:
                return i
    return -1


def to_int(v):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def parse_meters(spec, meters):
    """米数优先；为空时尝试从规格（如 34×15m）解析"""
    if meters is not None:
        return meters
    if not spec:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*m\b", spec)
    if m:
        return int(float(m.group(1)))
    return None


def main():
    rows = read_rows()
    if not rows:
        print("模板为空，未生成任何数据")
        return 1
    headers = [str(h).strip() if h else "" for h in rows[0]]
    body = rows[1:]

    i_type = col_index(headers, "类型")
    i_name = col_index(headers, "纸张名称", "纸张", "品名")
    i_id = col_index(headers, "纸张ID", "ID", "编号")
    i_qty = col_index(headers, "数量", "张数")
    i_mi = col_index(headers, "米数")
    i_spec = col_index(headers, "尺寸", "规格")

    items = []
    for r in body:
        name = str(r[i_name]).strip() if i_name >= 0 and r[i_name] else ""
        if not name:
            continue
        typ = str(r[i_type]).strip() if i_type >= 0 and r[i_type] else ""
        pid = str(r[i_id]).strip() if i_id >= 0 and r[i_id] else ""
        qty = to_int(r[i_qty]) if i_qty >= 0 else None
        mi = to_int(r[i_mi]) if i_mi >= 0 else None
        spec = str(r[i_spec]).strip() if i_spec >= 0 and r[i_spec] else ""
        is_long = (typ == "长卷")
        m = parse_meters(spec, mi)
        total = m if is_long else qty
        items.append({
            "id": pid,
            "type": typ or "刀纸",
            "name": name,
            "qty": qty,
            "meters": m,
            "spec": spec,
            "total": total,
            "remaining": total,
        })

    data = {"updatedAt": date.today().isoformat(), "items": items}
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # 统计
    from collections import Counter
    cnt = Counter(i["type"] for i in items)
    lv = sum(i["meters"] or 0 for i in items if i["type"] == "长卷")
    sheets = sum(i["qty"] or 0 for i in items if i["type"] in ("刀纸", "条屏"))
    print(f"✓ 已生成 {JSON_OUT}")
    print(f"  共 {len(items)} 条：{dict(cnt)}")
    print(f"  长卷总米数 {lv} m · 纸品总张数 {sheets} 张")
    print(f"  更新日期 {data['updatedAt']}")

    if "--open" in sys.argv:
        import webbrowser
        webbrowser.open("http://localhost:8777/shufa-inventory.html")
    return 0


if __name__ == "__main__":
    sys.exit(main())
